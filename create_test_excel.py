from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Add sheets with names
wb.create_sheet("Sheet1")
wb.create_sheet("Sheet2")
wb.create_sheet("Sheet3")

# Remove the default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# Save the workbook to a file
wb.save("test.xlsx")
print("Sample Excel file 'test.xlsx' created successfully.")