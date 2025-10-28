import tempfile
import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parent.parent))  # Add backend directory to PYTHONPATH

from parsers.excel_parser import extract_sheet_names
import openpyxl

def test_extract_sheet_names():
    # Create a temporary Excel file with multiple sheets
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        workbook.create_sheet("Sheet1")
        workbook.create_sheet("Sheet2")
        workbook.create_sheet("Sheet3")
        workbook.save(tmp.name)

        # Test the extract_sheet_names function
        sheet_names = extract_sheet_names(tmp.name)
        assert sheet_names == ["Sheet1", "Sheet2", "Sheet3"], f"Unexpected sheet names: {sheet_names}"

    print("Test passed: extract_sheet_names works correctly.")

if __name__ == "__main__":
    test_extract_sheet_names()