import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Union
import logging
import traceback
from datetime import datetime, timedelta
from difflib import SequenceMatcher

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class QuestionnaireRow:
    sdg: Optional[str] = None
    sdg_target_detailed: Optional[str] = None
    sdg_target: Optional[str] = None
    sustainability_dimension: Optional[str] = None
    kpi: Optional[str] = None
    question: Optional[str] = None
    scoring: Optional[str] = None
    source: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    comment: Optional[str] = None

@dataclass
class QuestionnaireData:
    textile: List[QuestionnaireRow] = field(default_factory=list)
    fertilizer: List[QuestionnaireRow] = field(default_factory=list)

class QuestionnaireParser:
    def __init__(self, file_path: str, sheet_names: List[str] = ["Textile_revised", "Fertilizer_revised"]):
        self.file_path = file_path
        self.sheet_names = sheet_names
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            logger.info(f"Successfully loaded Excel file: {file_path}")
        except Exception as e:
            logger.error(f"Failed to load workbook: {e}")
            raise

    def normalize_key(self, key: Optional[str]) -> Optional[str]:
        if key:
            normalized = key.strip().lower()
            normalized = re.sub(r'[^a-z0-9]', '_', normalized)
            normalized = re.sub(r'_+', '_', normalized).strip('_')
            return normalized
        return None

    def extract_questionnaire_data(self, sheet_name: str) -> List[dict]:
        if sheet_name not in self.wb.sheetnames:
            logger.warning(f"Sheet {sheet_name} not found")
            return []

        ws = self.wb[sheet_name]
        rows = []

        # Find header row (assuming row 1)
        header_row = 1
        headers = [self.normalize_key(cell.value) for cell in ws[header_row] if cell.value]

        logger.debug(f"Headers for {sheet_name}: {headers}")

        # Extract data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_idx]]
            if all(v is None for v in row_values):
                continue  # Skip empty rows

            row_dict = {}
            for idx, header in enumerate(headers):
                if idx < len(row_values):
                    row_dict[header] = row_values[idx]

            rows.append(row_dict)

        logger.debug(f"Extracted {len(rows)} rows from {sheet_name}")
        return rows

    def parse_all_data(self) -> Dict[str, List[dict]]:
        try:
            data = {
                'textile': self.extract_questionnaire_data("Textile_revised"),
                'fertilizer': self.extract_questionnaire_data("Fertilizer_revised")
            }
            return data
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise

def parse_excel_questionnaire(file_path: str) -> Dict[str, List[dict]]:
    try:
        parser = QuestionnaireParser(file_path)
        return parser.parse_all_data()
    except Exception as e:
        logger.error(f"Error in parse_excel_questionnaire: {e}\n{traceback.format_exc()}")
        raise

if __name__ == "__main__":
    file_path = "backend/data/IS_Questionnaires_revised_KK.xlsx"  # Replace with actual path
    try:
        parsed_data = parse_excel_questionnaire(file_path)
        print("Parsed Data:")
        print("Textile Rows:", len(parsed_data['textile']))
        print("Fertilizer Rows:", len(parsed_data['fertilizer']))
    except Exception as e:
        print(f"Error: {e}")