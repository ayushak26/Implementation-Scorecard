# questionnaire_parser.py

import openpyxl
import re
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Optional, Tuple, Any
import logging
import traceback

# -------------------- Logging --------------------
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s"
)
logger = logging.getLogger("QuestionnaireParser")

# -------------------- Data Models --------------------
@dataclass
class QuestionnaireRow:
    sdg_number: Optional[int] = None
    sdg: Optional[str] = None
    sector: Optional[str] = None
    sdg_target_detailed: Optional[str] = None
    sdg_target: Optional[str] = None
    sustainability_dimension: Optional[str] = None
    kpi: Optional[str] = None
    question: Optional[str] = None
    scoring_raw: Optional[str] = None
    scoring_clean: List[str] = field(default_factory=list)
    source: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None
    comment: Optional[str] = None


# -------------------- Parser --------------------
class QuestionnaireParser:
    """
    Sheet expectations:
      - Sector title cell at C3 (used as default/fallback).
      - Column B has SDG markers at exact rows:
           B1:'SDG' (title),
           B2:'SDG 1', B8:'SDG 2', B14:'SDG 3', ..., B98:'SDG 17'
      - A header row exists for the questionnaire table; we auto-detect it
        and map required headers (case-insensitive).
    """

    # Header variants for robust mapping
    REQUIRED_HEADERS: Dict[str, List[str]] = {
        "sdg": ["sdg"],
        "sdg_target_detailed": ["sdg target detailed", "sdg target (detailed)"],
        "sdg_target": ["sdg target", "sdg target (short)"],
        "sustainability_dimension": ["sustainability dimension", "dimension"],
        "kpi": ["kpi", "indicator", "metric"],
        "question": ["question", "assessment question", "assessment"],
        "scoring": ["scoring", "scores", "score"],
        "source": ["source", "reference"],
        "notes": ["notes", "note"],
        "status": ["status"],
        "comment": ["comment", "comments"],
    }

    # Exact SDG marker rows in column B (1-indexed)
    SDG_MARKERS: Dict[int, int] = {
        1:  2,  2:  8,  3: 14,  4: 20,  5: 26,  6: 32,
        7: 38,  8: 44,  9: 50, 10: 56, 11: 62, 12: 68,
        13: 74, 14: 80, 15: 86, 16: 92, 17: 98
    }
    SDG_TITLE_ROW = 1  # B1 should say "SDG"

    def __init__(self, file_path: str, sheet_names: Optional[List[str]] = None):
        self.file_path = file_path
        self.sheet_names = sheet_names or ["Textile_revised", "Fertilizer_revised"]
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
            logger.info(f"Loaded Excel: {file_path} | Sheets: {self.wb.sheetnames}")
        except Exception as e:
            logger.exception(f"Failed to load workbook: {e}")
            raise

    # -------------------- Utilities --------------------
    @staticmethod
    def _norm(v: Any) -> Optional[str]:
        if v is None:
            return None
        s = str(v).strip()
        return re.sub(r"\s+", " ", s)

    @staticmethod
    def _norm_key(s: Optional[str]) -> Optional[str]:
        if s is None:
            return None
        s = str(s).strip().lower()
        s = re.sub(r"[^a-z0-9]+", "_", s)
        return s.strip("_")

    def _detect_header_row_and_map(self, ws) -> Tuple[int, Dict[str, int]]:
        """
        Scan first 30 rows to find the most likely header row and map required headers to column indexes (0-based).
        """
        max_scan_rows = min(30, ws.max_row)
        best_row = None
        best_hit = -1
        best_map: Dict[str, int] = {}

        variants = {k: set([v.lower() for v in vals]) for k, vals in self.REQUIRED_HEADERS.items()}

        for r in range(1, max_scan_rows + 1):
            vals = [self._norm(c.value) for c in ws[r]]
            if all(v is None for v in vals):
                continue

            cand_map: Dict[str, int] = {}
            for idx, val in enumerate(vals):
                if not val:
                    continue
                low = val.lower()
                for k, alts in variants.items():
                    if low in alts and k not in cand_map:
                        cand_map[k] = idx

            hits = len(cand_map)
            if hits > best_hit:
                best_hit = hits
                best_row = r
                best_map = cand_map

        if best_row is None or best_hit == 0:
            logger.error("Could not detect a header row with required columns. Check your sheet headers.")
            raise ValueError("Header row not found")

        missing = [k for k in self.REQUIRED_HEADERS if k not in best_map]
        if missing:
            logger.warning(f"Header row detected at R{best_row}; missing headers: {missing}")
        else:
            logger.info(f"Header row detected at R{best_row}; all required headers mapped.")

        # Detailed map logging
        for k, col in best_map.items():
            addr = f"{openpyxl.utils.get_column_letter(col+1)}{best_row}"
            logger.debug(f"Header map: {k} -> {addr} ('{ws.cell(best_row, col+1).value}')")

        return best_row, best_map

    def _build_sdg_ranges(self, ws) -> Dict[int, Tuple[int, int]]:
        """
        Build inclusive row ranges per SDG using fixed marker rows in column B.
        """
        total_rows = ws.max_row
        ranges: Dict[int, Tuple[int, int]] = {}

        for sdg in range(1, 18):
            start = self.SDG_MARKERS[sdg]
            end = (self.SDG_MARKERS[sdg + 1] - 1) if sdg < 17 else total_rows
            ranges[sdg] = (start, end)

        # Verify the marker text in column B and log ranges
        b1 = self._norm(ws.cell(row=self.SDG_TITLE_ROW, column=2).value)
        if not b1 or "sdg" not in b1.lower():
            logger.warning(f"B1 expected to contain 'SDG', found: '{b1}'")
        for sdg, (r1, r2) in ranges.items():
            label = self._norm(ws.cell(row=r1, column=2).value)
            logger.debug(f"SDG {sdg} block: B{r1}='{label}' rows={r1}-{r2}")

        return ranges

    def _extract_sector_default(self, ws) -> Optional[str]:
        try:
            sector = self._norm(ws["C3"].value)
        except Exception:
            sector = None
        logger.info(f"Default sector from C3: {sector}")
        return sector

    def _extract_sector_by_sdg(self, ws, sdg_ranges: Dict[int, Tuple[int, int]]) -> Dict[int, Optional[str]]:
        """
        For each SDG, use column C at the SDG's marker row; fallback to C3 if empty.
        """
        default_sector = self._extract_sector_default(ws)
        by_sdg: Dict[int, Optional[str]] = {}
        for sdg, (start_row, _) in sdg_ranges.items():
            v = self._norm(ws.cell(row=start_row, column=3).value)  # C at marker row
            by_sdg[sdg] = v if v else default_sector
            logger.debug(f"SDG {sdg} sector: {by_sdg[sdg]!r} (C{start_row} or fallback C3)")
        return by_sdg

    @staticmethod
    def _parse_scoring(scoring_val: Optional[str]) -> List[str]:
        """
        Clean scoring: split by ';' or newline and remove leading numbering like '1 -', '2:', '3)'.
        """
        if not scoring_val:
            return []
        txt = str(scoring_val)
        parts = re.split(r"[;\n]+", txt)
        out: List[str] = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            p = re.sub(r"^\s*\(?\d+\)?\s*[:\-–\.]\s*", "", p)
            out.append(p.strip())
        if not out:
            out = [re.sub(r"^\s*\(?\d+\)?\s*[:\-–\.]\s*", "", txt).strip()]
        return out

    def _sheet_rows(
        self,
        ws,
        header_row: int,
        col_map: Dict[str, int],
        sdg_ranges: Dict[int, Tuple[int, int]],
        sector_by_sdg: Dict[int, Optional[str]],
    ) -> List[Dict]:
        """
        Read rows under header, assign SDG by row range, and extract only mapped columns.
        """
        records: List[Dict] = []
        data_start = header_row + 1
        data_end = ws.max_row

        # Build row -> sdg_number lookup
        row_to_sdg: Dict[int, int] = {}
        for sdg, (r1, r2) in sdg_ranges.items():
            for r in range(r1, r2 + 1):
                row_to_sdg[r] = sdg

        # Iterate data rows
        for r in range(data_start, data_end + 1):
            row_vals = [c.value for c in ws[r]]
            if all(v is None for v in row_vals):
                continue

            sdg_number = row_to_sdg.get(r)
            if sdg_number is None:
                # Outside SDG ranges — skip
                continue

            def get(key: str) -> Optional[str]:
                if key not in col_map:
                    return None
                cidx = col_map[key] + 1  # to 1-based
                return self._norm(ws.cell(row=r, column=cidx).value)

            sector = sector_by_sdg.get(sdg_number)
            row = QuestionnaireRow(
                sdg_number=sdg_number,
                sdg=get("sdg") or f"SDG {sdg_number}",
                sector=sector,
                sdg_target_detailed=get("sdg_target_detailed"),
                sdg_target=get("sdg_target"),
                sustainability_dimension=get("sustainability_dimension"),
                kpi=get("kpi"),
                question=get("question"),
                scoring_raw=get("scoring"),
                scoring_clean=self._parse_scoring(get("scoring")),
                source=get("source"),
                notes=get("notes"),
                status=get("status"),
                comment=get("comment"),
            )

            # Skip rows that look like section headers (no actual content besides SDG label)
            core = [
                row.sdg_target_detailed, row.sdg_target, row.sustainability_dimension, row.kpi,
                row.question, row.scoring_raw, row.source, row.notes, row.status, row.comment
            ]
            if all(v in (None, "") for v in core):
                continue

            # Log first few per SDG for traceability
            if len([x for x in records if x.get("sdg_number") == sdg_number]) < 3:
                logger.debug(f"{ws.title}: R{r} SDG{sdg_number} sector='{sector}' KPI='{row.kpi}' Q='{(row.question or '')[:50]}'")

            records.append(asdict(row))

        logger.info(f"{ws.title}: collected {len(records)} questionnaire rows")
        return records

    # -------------------- Public API --------------------
    def extract_questionnaire_data(self, sheet_name: str) -> Dict[str, Any]:
        if sheet_name not in self.wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found; skipping.")
            return {"rows": [], "sector_by_sdg": {}}

        ws = self.wb[sheet_name]

        # 1) SDG ranges from fixed markers
        sdg_ranges = self._build_sdg_ranges(ws)

        # 2) Sector per SDG (C at marker row; fallback C3)
        sector_by_sdg = self._extract_sector_by_sdg(ws, sdg_ranges)

        # 3) Detect header row & column mapping
        header_row, col_map = self._detect_header_row_and_map(ws)
        logger.debug(f"{sheet_name}: header at R{header_row}, map={col_map}")

        # 4) Extract rows
        rows = self._sheet_rows(ws, header_row, col_map, sdg_ranges, sector_by_sdg)

        # Put a meta snapshot in the log for quick visibility
        logger.debug(f"{sheet_name}: sector_by_sdg = {sector_by_sdg}")

        return {"rows": rows, "sector_by_sdg": sector_by_sdg}

    def parse_all_data(self) -> Dict[str, Dict[str, Any]]:
        try:
            out: Dict[str, Dict[str, Any]] = {}
            for sheet in self.sheet_names:
                logger.info(f"---- Parsing sheet: {sheet} ----")
                key = self._norm_key(sheet) or sheet
                out[key] = self.extract_questionnaire_data(sheet)
            return out
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}\n{traceback.format_exc()}")
            raise


# -------------------- Convenience wrapper --------------------
def parse_excel_questionnaire(file_path: str, sheet_names: Optional[List[str]] = None) -> Dict[str, Dict[str, Any]]:
    parser = QuestionnaireParser(file_path, sheet_names)
    return parser.parse_all_data()


# -------------------- CLI --------------------
if __name__ == "__main__":
    # Replace with your actual file path
    xlsx = "backend/data/IS_Questionnaires_revised_KK.xlsx"
    try:
        parsed = parse_excel_questionnaire(xlsx)
        for sheet, payload in parsed.items():
            rows = payload.get("rows", [])
            sector_by_sdg = payload.get("sector_by_sdg", {})
            logger.info(f"[{sheet}] rows={len(rows)} | sectors={sector_by_sdg}")
            # Peek first example from each SDG
            seen = set()
            for rec in rows:
                s = rec["sdg_number"]
                if s not in seen:
                    seen.add(s)
                    logger.debug(f"SDG {s} sector={rec['sector']} KPI='{rec.get('kpi')}' Q='{(rec.get('question') or '')[:60]}'")
    except Exception as e:
        logger.exception(f"CLI error: {e}")
