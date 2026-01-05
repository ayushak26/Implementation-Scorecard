# questionnaire_parser.py

import logging
import re
from dataclasses import dataclass, asdict, field
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple, Union
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

# ============================== Logging ==============================
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
)
logger = logging.getLogger("QuestionnaireParser")

# ============================== Constants ==============================
SDG_DESCRIPTIONS: Dict[int, str] = {
    1: "No Poverty",
    2: "Zero Hunger",
    3: "Good Health & Well-being",
    4: "Ensure inclusive and equitable quality education and promote lifelong learning opportunities for all",
    5: "Gender Equality",
    6: "Clean Water & Sanitation",
    7: "Affordable and Clean Energy",
    8: "Decent Work & Economic Growth",
    9: "Build resilient infrastructure, promote inclusive and sustainable industrialization and foster innovation",
    10: "Reduce inequality within and among countries",
    11: "Make cities and human settlements inclusive, safe, resilient and sustainable",
    12: "Ensure sustainable consumption and production patterns",
    13: "Climate Action",
    14: "Life Below Water",
    15: "Life on Land",
    16: "Peace, Justice and Strong Institutions",
    17: "Partnerships for the Goals",
}

ALLOWED_SECTORS_CANON = {"Textiles", "Fertilizers", "Packaging"}

SECTOR_SYNONYMS: Dict[str, str] = {
    # Textiles
    "textile": "Textiles",
    "textiles": "Textiles",
    "textil": "Textiles",
    "fabric": "Textiles",
    "garment": "Textiles",
    "apparel": "Textiles",
    # Fertilizers
    "fertilizer": "Fertilizers",
    "fertilizers": "Fertilizers",
    "fert": "Fertilizers",
    # Packaging
    "packaging": "Packaging",
    "package": "Packaging",
    "packing": "Packaging",
    "pack": "Packaging",
}

RUBRIC_CANON: Dict[int, str] = {
    0: "N/A",
    1: "Issue identified, but no plans for further actions",
    2: "Issue identified, starts planning further actions",
    3: "Action plan with clear targets and deadlines in place",
    4: "Action plan operational - some progress in established targets",
    5: "Action plan operational - achieving the target set",
}

RUBRIC_PHRASES: Dict[int, List[str]] = {
    0: ["n/a", "na", "not applicable"],
    1: ["issue identified", "no plans"],
    2: ["starts planning", "planning further actions"],
    3: ["action plan", "clear targets", "deadlines"],
    4: ["operational", "some progress"],
    5: ["operational", "achieving the target", "achieving target"],
}

# ---- IMPORTANT: SDG markers are now in COLUMN C at the given rows ----
SDG_MARKERS_C: Dict[int, int] = {
    1: 2,
    2: 16,
    3: 30,
    4: 44,
    5: 58,
    6: 72,
    7: 86,
    8: 102,
    9: 116,
    10: 130,
    11: 144,
    12: 158,
    13: 172,
    14: 186,
    15: 200,
    16: 214,
    17: 228,
}
SDG_HEADER_COL = 3  # Column C
MAX_ROWS = 250      # per your constraint

# ============================== Recommendation Helpers ==============================
MATURITY_KEY_MAP = {
    "awareness": "awareness",
    "aware": "awareness",
    "developing": "developing",
    "develop": "developing",
    "leading": "leading",
    "lead": "leading",
}

def _norm_maturity(v: Optional[str]) -> Optional[str]:
    if not v:
        return None
    s = re.sub(r"\s+", " ", str(v).strip().lower())
    for k, canon in MATURITY_KEY_MAP.items():
        if s == k or k in s:
            return canon
    return None


# ============================== Data Model ==============================
@dataclass
class QuestionnaireRow:
    sdg_number: Optional[int] = None
    sdg_description: Optional[str] = None
    sector: Optional[str] = None
    sdg_target: Optional[str] = None
    sustainability_dimension: Optional[str] = None
    kpi: Optional[str] = None
    question: Optional[str] = None
    score: Optional[int] = None
    score_description: Optional[str] = None
    source: Optional[str] = None
    notes: Optional[str] = None

    # grouped recommendations per maturity level
    # example:
    # {
    #   "awareness": {"text": "...", "source": "..."},
    #   "developing": {"text": "...", "source": "..."},
    #   "leading": {"text": "...", "source": "..."},
    # }
    recommendations: Dict[str, Dict[str, Optional[str]]] = field(default_factory=dict)


# ============================== Parser ==============================
class QuestionnaireParser:
    """
    Parses questionnaire rows and groups recommendations.

    Behavior:
      - SDG ranges are determined from fixed SDG headers in column C (C2, C16, ... C228)
      - Data rows are assigned to SDGs based on those ranges (up to MAX_ROWS)
      - Recommendations are grouped across multiple rows under ONE question record using maturity level.
    """

    REQUIRED_HEADERS: Dict[str, List[str]] = {
        "sdg_target": ["sdg target", "sdg target (short)"],
        "sustainability_dimension": ["sustainability dimension", "dimension"],
        "kpi": ["kpi", "indicator", "metric"],
        "question": ["question", "assessment question", "assessment"],
        "scoring": ["scoring", "scores", "score"],
        "source": ["source", "reference"],  # first "Source"
        "notes": ["notes", "note"],

        # recommendations
        "maturity_level": ["maturity level", "maturity"],
        "recommendations": ["recommendations", "recommendation"],
        # second "Source" (recommendation source) is detected dynamically
    }

    def __init__(self, file_source: Union[str, BytesIO], sheet_names: Optional[List[str]] = None):
        self.file_source = file_source
        self._requested_sheets = sheet_names or [
            "Textiles",
            "Fertilizers",
            "Packaging",
            "Textile_revised",
            "Fertilizer_revised",
            "Packaging_revised",
        ]

        try:
            if isinstance(file_source, BytesIO):
                file_source.seek(0)
                self.wb = openpyxl.load_workbook(file_source, data_only=True)
                logger.info(f"Loaded Excel from BytesIO | Sheets: {self.wb.sheetnames}")
            else:
                self.wb = openpyxl.load_workbook(file_source, data_only=True)
                logger.info(f"Loaded Excel: {file_source} | Sheets: {self.wb.sheetnames}")
        except Exception as e:
            logger.exception(f"Failed to load workbook: {e}")
            raise

        self.sheet_names = self._normalize_sheet_list(self._requested_sheets)
        logger.info(f"Using sheets: {self.sheet_names}")

    # -------------------- helpers: normalization --------------------
    @staticmethod
    def _norm(v: Any) -> Optional[str]:
        if v is None:
            return None
        s = re.sub(r"\s+", " ", str(v).strip())
        return s if s else None

    @staticmethod
    def _norm_key(s: Optional[str]) -> Optional[str]:
        if s is None:
            return None
        s = re.sub(r"[^a-z0-9]+", "_", str(s).strip().lower()).strip("_")
        return s or None

    # -------------------- helpers: sheet resolution --------------------
    def _resolve_sheet_name(self, desired: str) -> Optional[str]:
        if not desired:
            return None
        try:
            low = desired.strip().lower()
        except (AttributeError, TypeError):
            return None

        avail = self.wb.sheetnames

        for a in avail:
            if a.lower() == low:
                return a

        best_match, best_score = None, 0.0
        for a in avail:
            score = SequenceMatcher(None, low, a.lower()).ratio()
            if score > best_score:
                best_score = score
                best_match = a

        if best_score > 0.6:
            return best_match
        return None

    def _normalize_sheet_list(self, sheets: List[str]) -> List[str]:
        out: List[str] = []
        for s in sheets:
            resolved = self._resolve_sheet_name(s)
            if resolved:
                if resolved not in out:
                    out.append(resolved)
            else:
                logger.warning(f"Sheet '{s}' not found by name/fuzzy; skipping.")
        return out or self.wb.sheetnames

    # -------------------- headers --------------------
    def _detect_header_row_and_map(self, ws) -> Tuple[int, Dict[str, int], Optional[int]]:
        """
        Returns:
          header_row,
          col_map (0-based indices),
          rec_source_col_idx (0-based) for the SECOND "Source" column if present.
        """
        max_scan_rows = min(30, ws.max_row, MAX_ROWS)
        best_row, best_hit, best_map = None, -1, {}
        variants = {k: {v.lower() for v in vals} for k, vals in self.REQUIRED_HEADERS.items()}
        best_rec_source_idx: Optional[int] = None

        for r in range(1, max_scan_rows + 1):
            vals = [self._norm(c.value) for c in ws[r]]
            if all(v is None for v in vals):
                continue

            cand_map: Dict[str, int] = {}
            source_indices: List[int] = []

            for idx, val in enumerate(vals):
                if not val:
                    continue
                low = val.strip().lower()

                if low in {"source", "reference"}:
                    source_indices.append(idx)

                for k, alts in variants.items():
                    if low in alts and k not in cand_map:
                        cand_map[k] = idx

            hits = len(cand_map)
            if hits > best_hit:
                best_row, best_hit, best_map = r, hits, cand_map
                best_rec_source_idx = source_indices[1] if len(source_indices) >= 2 else None

        if best_row is None or best_hit == 0:
            logger.error("Could not detect a header row with required columns. Check your sheet headers.")
            raise ValueError("Header row not found")

        for k, col in best_map.items():
            addr = f"{get_column_letter(col+1)}{best_row}"
            logger.debug(f"Header map: {k} -> {addr} ('{ws.cell(best_row, col+1).value}')")

        if best_rec_source_idx is not None:
            addr = f"{get_column_letter(best_rec_source_idx+1)}{best_row}"
            logger.info(
                f"Detected recommendation source column at {addr} "
                f"('{ws.cell(best_row, best_rec_source_idx+1).value}')"
            )

        missing = [k for k in self.REQUIRED_HEADERS if k not in best_map]
        if missing:
            logger.warning(f"Header detected at R{best_row}; missing headers: {missing}")
        else:
            logger.info(f"Header detected at R{best_row}; all required headers mapped.")

        return best_row, best_map, best_rec_source_idx

    # -------------------- SDG ranges (FIXED COLUMN C + FIXED ROWS) --------------------
    def _build_sdg_ranges(self, ws) -> Dict[int, Tuple[int, int]]:
        """
        Builds SDG ranges using the exact mapping provided:
          SDG 1 header at C2, SDG 2 at C16, ... SDG 17 at C228.
        End row is the row before the next SDG header row.
        Final SDG ends at min(ws.max_row, MAX_ROWS).
        """
        total_rows = min(ws.max_row, MAX_ROWS)

        ranges: Dict[int, Tuple[int, int]] = {}
        for sdg in range(1, 18):
            start = SDG_MARKERS_C[sdg]
            end = (SDG_MARKERS_C[sdg + 1] - 1) if sdg < 17 else total_rows
            ranges[sdg] = (start, end)

        # debug logs to verify what is in column C at each marker row
        for sdg, (r1, r2) in ranges.items():
            header_label = self._norm(ws.cell(row=r1, column=SDG_HEADER_COL).value)
            logger.debug(
                f"{ws.title}: SDG {sdg} header C{r1}='{header_label}' rows={r1}-{r2}"
            )

        return ranges

    # -------------------- sector helpers --------------------
    def _canonicalize_sector(self, raw: Optional[str]) -> Optional[str]:
        if not raw:
            return None
        s = re.sub(r"\bsector\b\s*[:\-–|]?\s*", "", str(raw), flags=re.IGNORECASE)
        tokens = re.split(r"[,/;|]+", s)
        candidates: List[str] = []

        for t in tokens if tokens else [s]:
            tt = re.sub(r"\bsector\b", "", t, flags=re.IGNORECASE).strip()
            if not tt:
                continue
            low = tt.lower()

            if low in SECTOR_SYNONYMS:
                candidates.append(SECTOR_SYNONYMS[low])
                continue

            best_match, best_ratio = None, 0.0
            for key, canon in SECTOR_SYNONYMS.items():
                ratio = SequenceMatcher(None, low, key).ratio()
                if ratio > best_ratio:
                    best_match, best_ratio = canon, ratio
            if best_match and best_ratio >= 0.80:
                candidates.append(best_match)

        uniq: List[str] = []
        for c in candidates:
            if c in ALLOWED_SECTORS_CANON and c not in uniq:
                uniq.append(c)

        if not uniq:
            return None
        if len(uniq) > 1:
            logger.warning(f"Multiple sector candidates {uniq}; selecting '{uniq[0]}'")
        return uniq[0]

    def _extract_sector_default(self, ws, sheet_name: Optional[str]) -> Optional[str]:
        # Keep your existing behavior (C3 if present), else infer from sheet name
        try:
            raw = self._norm(ws["C3"].value)
        except Exception:
            raw = None

        canon = self._canonicalize_sector(raw)

        if not canon and sheet_name:
            name = sheet_name.lower()
            if "textile" in name:
                canon = "Textiles"
            elif "fertilizer" in name or "fertilis" in name:
                canon = "Fertilizers"
            elif "packag" in name:
                canon = "Packaging"

        logger.info(f"Default sector resolved: {canon!r} (C3 raw={raw!r}, sheet_name={sheet_name!r})")
        return canon

    def _extract_sector_by_sdg(
        self, ws, sdg_ranges: Dict[int, Tuple[int, int]], sheet_name: Optional[str]
    ) -> Dict[int, Optional[str]]:
        """
        Your SDG headers are in column C. Sector is usually in C3 / sheet name,
        but we also keep the same per-SDG lookup logic if you later add sector per SDG.
        """
        default_sector = self._extract_sector_default(ws, sheet_name)
        by_sdg: Dict[int, Optional[str]] = {}

        for sdg, (start_row, _) in sdg_ranges.items():
            # If you ever put sector on the SDG header row, define which column it's in.
            # For now we keep it consistent with earlier parser behavior:
            # attempt column C at start_row (same col), then fallback to default.
            raw = self._norm(ws.cell(row=start_row, column=SDG_HEADER_COL).value)
            # NOTE: raw is SDG header text, so canonicalize_sector likely returns None, hence fallback
            canon = self._canonicalize_sector(raw) or default_sector
            by_sdg[sdg] = canon
            logger.debug(f"[{ws.title}] SDG {sdg}: C{start_row} raw={raw!r} -> sector={canon!r} (fallback={default_sector!r})")

        return by_sdg

    # -------------------- scoring helpers --------------------
    @staticmethod
    def _clean_scoring_text(txt: str) -> str:
        return re.sub(r"^\s*\(?\d+\)?\s*[:\-–\.]\s*", "", txt).strip()

    def _extract_score_number(self, scoring_val: Optional[str]) -> Optional[int]:
        if not scoring_val:
            return None
        txt = str(scoring_val).strip()

        m = re.match(r"^\s*\(?([0-5])\)?(?:\s*[:\-\–\.\)]|\s)", txt)
        if m:
            return int(m.group(1))

        nums = re.findall(r"(?<!\d)([0-5])\s*[:\-\–\.\)]", txt)
        uniq = sorted({int(n) for n in nums})
        if len(uniq) == 1:
            return uniq[0]
        if len(uniq) > 1:
            return None

        low = re.sub(r"\s+", " ", txt).lower()
        if re.search(r"\b(n/?a|not applicable)\b", low):
            return 0

        def phrase_score(targets: List[str]) -> int:
            s = 0
            for ph in targets:
                if ph in low:
                    s += 2
                else:
                    r = SequenceMatcher(None, ph, low).ratio()
                    if r > 0.6:
                        s += 1
            return s

        best_num, best_s = None, -1
        for num, phrases in RUBRIC_PHRASES.items():
            ps = phrase_score(phrases)
            if ps > best_s:
                best_num, best_s = num, ps
        return best_num if best_s >= 2 else None

    def _derive_score_description(self, scoring_cell: Optional[str], score: Optional[int]) -> Optional[str]:
        if score is not None:
            return RUBRIC_CANON.get(score)
        if scoring_cell:
            return self._clean_scoring_text(str(scoring_cell))
        return None

    # -------------------- row extraction (grouped recommendations) --------------------
    def _sheet_rows(
        self,
        ws,
        header_row: int,
        col_map: Dict[str, int],
        rec_source_col_idx: Optional[int],
        sdg_ranges: Dict[int, Tuple[int, int]],
        sector_by_sdg: Dict[int, Optional[str]],
    ) -> List[Dict]:
        records: List[Dict] = []
        data_start = header_row + 1
        data_end = min(ws.max_row, MAX_ROWS)

        # Map each row to an SDG number based on fixed ranges
        row_to_sdg: Dict[int, int] = {}
        for sdg, (r1, r2) in sdg_ranges.items():
            for r in range(r1, min(r2, data_end) + 1):
                row_to_sdg[r] = sdg

        def get_cell_str(r: int, key: str) -> Optional[str]:
            if key not in col_map:
                return None
            cidx = col_map[key] + 1  # to 1-based
            return self._norm(ws.cell(row=r, column=cidx).value)

        def get_rec_source(r: int) -> Optional[str]:
            if rec_source_col_idx is None:
                return None
            return self._norm(ws.cell(row=r, column=rec_source_col_idx + 1).value)

        def has_any_question_metadata(r: int) -> bool:
            # identifies the first row of a question group
            return any(
                get_cell_str(r, k)
                for k in ("sdg_target", "sustainability_dimension", "kpi", "question")
                if k in col_map
            )

        current: Optional[QuestionnaireRow] = None

        def flush_current():
            nonlocal current
            if not current:
                return

            # Strict skip: if all detail fields are empty AND no recommendations, drop it
            empties = [
                current.sustainability_dimension,
                current.kpi,
                current.question,
                current.score,
                current.score_description,
                current.source,
                current.notes,
            ]
            if all(v in (None, "", []) for v in empties) and not current.recommendations:
                current = None
                return

            records.append(asdict(current))
            current = None

        for r in range(data_start, data_end + 1):
            # ignore fully empty rows
            row_vals = [c.value for c in ws[r]]
            if all(v is None for v in row_vals):
                continue

            sdg_number = row_to_sdg.get(r)
            if sdg_number is None:
                continue

            sector = sector_by_sdg.get(sdg_number)

            # Read recommendation fields from this row
            maturity_raw = get_cell_str(r, "maturity_level")
            maturity_key = _norm_maturity(maturity_raw)
            rec_text = get_cell_str(r, "recommendations")
            rec_src = get_rec_source(r)

            # Start a new question group when metadata present
            if has_any_question_metadata(r):
                flush_current()

                scoring_cell = get_cell_str(r, "scoring")
                score = self._extract_score_number(scoring_cell)
                score_desc = self._derive_score_description(scoring_cell, score)

                current = QuestionnaireRow(
                    sdg_number=sdg_number,
                    sdg_description=SDG_DESCRIPTIONS.get(sdg_number),
                    sector=sector,
                    sdg_target=get_cell_str(r, "sdg_target"),
                    sustainability_dimension=get_cell_str(r, "sustainability_dimension"),
                    kpi=get_cell_str(r, "kpi"),
                    question=get_cell_str(r, "question"),
                    score=score,
                    score_description=score_desc,
                    source=get_cell_str(r, "source"),
                    notes=get_cell_str(r, "notes"),
                )

            # Attach rec rows (works for both the first row and follow-on rec-only rows)
            if current and maturity_key and (rec_text or rec_src):
                current.recommendations[maturity_key] = {
                    "text": rec_text,
                    "source": rec_src,
                }

            # Rec-only row but no current question context -> skip safely
            if (maturity_key or rec_text) and not current:
                logger.debug(f"{ws.title}: R{r} has recommendation info but no active question context; skipping attach.")
                continue

        flush_current()

        logger.info(f"{ws.title}: collected {len(records)} questionnaire question-records (grouped recommendations)")
        return records

    # -------------------- public API --------------------
    def extract_questionnaire_data(self, sheet_name: str) -> Dict[str, Any]:
        resolved = self._resolve_sheet_name(sheet_name)
        if not resolved:
            logger.warning(f"Sheet '{sheet_name}' not found; skipping.")
            return {"rows": [], "sector_by_sdg": {}}

        ws = self.wb[resolved]
        sdg_ranges = self._build_sdg_ranges(ws)
        sector_by_sdg = self._extract_sector_by_sdg(ws, sdg_ranges, resolved)
        header_row, col_map, rec_source_col_idx = self._detect_header_row_and_map(ws)

        logger.debug(
            f"{resolved}: header at R{header_row}, map={col_map}, rec_source_col_idx={rec_source_col_idx}"
        )

        rows = self._sheet_rows(ws, header_row, col_map, rec_source_col_idx, sdg_ranges, sector_by_sdg)
        return {"rows": rows, "sector_by_sdg": sector_by_sdg}

    def parse_all_data(self) -> Dict[str, Dict[str, Any]]:
        out: Dict[str, Dict[str, Any]] = {}
        for sheet in self.sheet_names:
            logger.info(f"---- Parsing sheet: {sheet} ----")
            key = self._norm_key(sheet) or sheet
            out[key] = self.extract_questionnaire_data(sheet)
        return out

    def close(self):
        if hasattr(self, "wb"):
            self.wb.close()


# ============================== Convenience Wrappers ==============================
def parse_excel_questionnaire(file_source: Union[str, BytesIO], sheet_names: Optional[List[str]] = None) -> Dict[str, Dict[str, Any]]:
    parser = QuestionnaireParser(file_source, sheet_names)
    try:
        return parser.parse_all_data()
    finally:
        parser.close()


def extract_questions_for_interactive(file_source: Union[str, BytesIO], sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Extract questions + grouped recommendations.
    """
    if sheet_name is None or sheet_name == "":
        sheets_to_process = ["Textiles", "Fertilizers", "Packaging"]
    else:
        sheets_to_process = [sheet_name]

    parser = QuestionnaireParser(file_source, sheets_to_process)

    all_questions = []
    last_sector = "General"

    try:
        for sheet in parser.sheet_names:
            data = parser.extract_questionnaire_data(sheet)

            for row in data.get("rows", []):
                if row.get("question"):
                    all_questions.append({
                        "id": row.get("id") or f"q_{len(all_questions) + 1}",
                        "sdg_number": row.get("sdg_number"),
                        "sdg_description": row.get("sdg_description"),
                        "sdg_target": row.get("sdg_target"),
                        "sustainability_dimension": row.get("sustainability_dimension"),
                        "kpi": row.get("kpi"),
                        "question": row.get("question"),
                        "sector": row.get("sector", "Unknown"),
                        "recommendations": row.get("recommendations", {}),
                    })
                    last_sector = row.get("sector", last_sector)
    finally:
        parser.close()

    return {"questions": all_questions, "sector": last_sector}


# ============================== CLI ==============================
if __name__ == "__main__":
    xlsx = "backend/data/final.xlsx"  # change if needed
    try:
        parsed = parse_excel_questionnaire(xlsx)
        print(f"Parsed data from {xlsx}:")
        for sheet, data in parsed.items():
            print(f"Sheet: {sheet} | Rows: {len(data['rows'])}")
    except Exception as e:
        logger.exception(f"CLI error: {e}")
